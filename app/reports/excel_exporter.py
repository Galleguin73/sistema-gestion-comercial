# Archivo: app/reports/excel_exporter.py

import openpyxl
from openpyxl.styles import Font, Alignment
from app.database import ventas_db, compras_db, clientes_db, proveedores_db, articulos_db
from collections import defaultdict

def _estilizar_cabecera(ws):
    """Aplica un estilo (negrita y centrado) a la primera fila de la hoja y autoajusta columnas."""
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        ws.column_dimensions[column].width = adjusted_width

# --- NUEVA FUNCIÓN AÑADIDA ---
def exportar_listado_articulos(path):
    try:
        articulos = articulos_db.obtener_articulos(incluir_inactivos=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado de Artículos"
        
        cabecera = ["ID", "Código Barras", "Marca", "Nombre", "Stock", "Precio Venta", "Estado", "Unidad"]
        ws.append(cabecera)

        for art in articulos:
            ws.append(list(art))
        
        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Listado de artículos exportado exitosamente."
    except Exception as e:
        return False, f"Error al exportar listado de artículos: {e}"

# --- NUEVA FUNCIÓN AÑADIDA ---
def exportar_listado_reposicion(path):
    try:
        articulos = articulos_db.obtener_articulos_stock_bajo()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado de Reposición"
        
        cabecera = ["Nombre", "Stock Actual", "Stock Mínimo"]
        ws.append(cabecera)
        
        for art in articulos:
            ws.append(list(art))

        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Listado de reposición exportado exitosamente."
    except Exception as e:
        return False, f"Error al exportar listado de reposición: {e}"

def exportar_ventas_periodo(path, fecha_desde, fecha_hasta):
    try:
        ventas = ventas_db.obtener_ventas_por_periodo(fecha_desde, fecha_hasta)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas por Período"
        cabecera = ["ID Venta", "Fecha", "Cliente", "Comprobante", "Monto Total", "Estado"]
        ws.append(cabecera)
        
        total_facturado = sum(v[4] for v in ventas if v[5] != 'ANULADA')
        for venta in ventas: ws.append(venta)
        
        ws.append([])
        total_row = ws.max_row
        ws.cell(row=total_row, column=4, value="Total Facturado:").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_facturado).number_format = '"$"#,##0.00'

        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Reporte de ventas exportado exitosamente."
    except Exception as e:
        return False, f"Error al exportar ventas: {e}"

def exportar_compras_periodo(path, fecha_desde, fecha_hasta):
    try:
        compras = compras_db.obtener_compras_por_periodo(fecha_desde, fecha_hasta)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compras por Período"
        cabecera = ["ID Compra", "Fecha", "Proveedor", "N° Factura", "Monto Total", "Estado"]
        ws.append(cabecera)
        for compra in compras: ws.append(compra)
        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Reporte de compras exportado exitosamente."
    except Exception as e:
        return False, f"Error al exportar compras: {e}"

def exportar_ventas_categorias(path, fecha_desde, fecha_hasta):
    try:
        datos = ventas_db.reporte_ventas_por_rubro_y_subrubro(fecha_desde, fecha_hasta)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas por Categoría"
        cabecera = ["Rubro", "Subrubro", "Cantidad Vendida", "Total Vendido (Neto)"]
        ws.append(cabecera)
        
        rubros_data = defaultdict(lambda: {'cantidad': 0, 'total': 0, 'subrubros': []})
        for rubro, subrubro, cantidad, total in datos:
            rubros_data[rubro]['cantidad'] += cantidad
            rubros_data[rubro]['total'] += total
            rubros_data[rubro]['subrubros'].append((subrubro, cantidad, total))
        
        for rubro, data in rubros_data.items():
            ws.append([rubro, "", data['cantidad'], data['total']])
            for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            for subrubro, cantidad, total in data['subrubros']:
                ws.append(["", f"  - {subrubro}", cantidad, total])
        
        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Reporte de categorías exportado exitosamente."
    except Exception as e:
        return False, f"Error al exportar categorías: {e}"

def exportar_ventas_articulo(path, fecha_desde, fecha_hasta):
    try:
        articulos = ventas_db.reporte_ventas_por_articulo(fecha_desde, fecha_hasta)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas por Artículo"
        cabecera = ["Código", "Artículo", "Marca", "Cant. Vendida", "Total Vendido"]
        ws.append(cabecera)
        for art in articulos: ws.append(art)
        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Reporte de ventas por artículo exportado."
    except Exception as e:
        return False, f"Error al exportar ventas por artículo: {e}"

def exportar_ventas_marca(path, fecha_desde, fecha_hasta):
    try:
        marcas = ventas_db.reporte_ventas_por_marca(fecha_desde, fecha_hasta)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas por Marca"
        cabecera = ["Marca", "Cant. Vendida", "Total Vendido"]
        ws.append(cabecera)
        for marca in marcas: ws.append(marca)
        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Reporte de ventas por marca exportado."
    except Exception as e:
        return False, f"Error al exportar ventas por marca: {e}"

def exportar_cc_cliente(path, cliente_id, fecha_desde, fecha_hasta):
    try:
        movimientos = clientes_db.obtener_cuenta_corriente_cliente(cliente_id, fecha_desde, fecha_hasta)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cta. Cte. Cliente"
        cabecera = ["Fecha", "Tipo Movimiento", "Monto", "Saldo Resultante"]
        ws.append(cabecera)
        for mov in movimientos: ws.append(mov)
        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Reporte de cuenta corriente de cliente exportado."
    except Exception as e:
        return False, f"Error al exportar cta. cte. de cliente: {e}"

def exportar_cc_proveedor(path, proveedor_id, fecha_desde, fecha_hasta):
    try:
        movimientos = proveedores_db.obtener_cuenta_corriente_proveedor(proveedor_id, fecha_desde, fecha_hasta)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cta. Cte. Proveedor"
        cabecera = ["Fecha", "Tipo Movimiento", "Monto", "Saldo Resultante"]
        ws.append(cabecera)
        for mov in movimientos: ws.append(mov)
        _estilizar_cabecera(ws)
        wb.save(path)
        return True, "Reporte de cuenta corriente de proveedor exportado."
    except Exception as e:
        return False, f"Error al exportar cta. cte. de proveedor: {e}"